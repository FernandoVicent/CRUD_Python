from random import sample
import openpyxl
from openpyxl import Workbook

planilha_estoque = openpyxl.Workbook()
print(planilha_estoque.sheetnames)

planilha_estoque.create_sheet('gerenciamento de estoque')

items_page = planilha_estoque['gerenciamento de estoque']
items_page.append(['Código','Nome do produto','Fabricante','Valor'])



print("Bem vindo ao gerenciamento de estoque feito por Fernando_Vicente")

estoque = []
peca = {}

def LeiaInt(msg):
    while True:
        try:
            n = int(input(msg))
        except (ValueError, TypeError):
            print('\033[31mERRO: por favor, Digite um opção valida.\033[m')
            continue
        except (KeyboardInterrupt):
            print('\n\033[32mUsuario preferiu nao digitar uma opção. \033[m')
            return 0
        else:
            return n
def Menu(lista):
    c = 1
    for item in lista:
        print(f'\033[33m {c} \033[m - \033[34m{item}\033[m')
        c += 1
    opc = LeiaInt('\033[32m Sua opção : \033[m ')
    return opc

def cadastrar_peca():
    main_numbers = sample(range(1, 100), 1)

    codigo = main_numbers[0]
    print(f"o código da sua peça é: {main_numbers[0]}")
    nome = input("Qual o nome da peça: ")
    fabricante = input("qual o fabricante da peça: ")
    valor = int(input("qual o valor da peça: "))
    peca = {'codigo':codigo, 'nome': nome, 'fabricante': fabricante, 'valor': valor}
    estoque.append(peca)
    print(peca['codigo'])
    print(peca['nome'])
    items_page.append([peca['codigo'],peca['nome'],peca['fabricante'],peca['valor']])
    planilha_estoque.save('planilha.xlsx')



def consultar_peca():
    consultas = Menu(['consultar todas as peças','consultar peças por codígo','consultar peças por fabricante','Retornar'])
    if consultas == 1:
        for s in estoque:
            print('_' * 20)
            print(f'código da peça: {s["codigo"]}\n'
                  f'peça: {s["nome"]}\n'
                  f'fabricante:{s["fabricante"]}\n'
                  f'R$: {s["valor"]}')
            print('_' * 20)


    elif consultas == 2:
        busca = int(input('qual o codigo que procuras: '))
        c = 0
        for y in estoque:
            c += 1
            if busca == y['codigo']:
                print('_' * 20)
                print(f'código da peça: {y["codigo"]}\n'
                      f'peça: {y["nome"]}\n'
                      f'fabricante:{y["fabricante"]}\n'
                      f'R$: {y["valor"]}')
                print('_' * 20)
            else:
                print('codigo não encontrado')
        # achado = next((x for x in estoque if x['codigo'] == busca),None)
        #
        # if achado == None:
        #     print('não encontramos esse codigo no nosso banco de dados')
        #     consultar_peca()
        #
        # print('_' * 20)
        # print(f'código da peça: {achado["codigo"]}\n'
        #       f'peça: {achado["nome"]}\n'
        #       f'fabricante:{achado["fabricante"]}\n'
        #       f'R$: {achado["valor"]}')
        # print('_' * 20)

    elif consultas == 3:
        busca_fab = input('qual o fabricante que procura: ')
        cont = 0
        for i in estoque:
            cont += 1
            if busca_fab == i['fabricante']:
                print('_' * 20)
                print(f'código da peça: {i["codigo"]}\n'
                      f'peça: {i["nome"]}\n'
                      f'fabricante:{i["fabricante"]}\n'
                      f'R$: {i["valor"]}')
                print('_' * 20)
            else:
                print('fabricante não encontrado no nosso banco de dados')

    elif consultas == 4:
        print('retornando')

def remover_peca():
    remover = int(input('digite o codigo da sua peça'))

    contador = 0
    for e in estoque:
        contador += 1
        if remover == e['codigo']:
            estoque.remove(e)



def inicio():
    while True:
        resposta = Menu(['Cadastrar Peça','Consultar Peça','remover peça','Sair'])
        if resposta == 1:
            cadastrar_peca()
        elif resposta == 2:
            consultar_peca()
        elif resposta ==3:
            remover_peca()
        elif resposta == 4:
            break
        else:
            print('ops não entendemos poderia digitar novamente')
            continue
inicio()
